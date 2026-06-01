from app.engines.knowledge_engine.parser.base import BaseDocumentParser


class ExcelParser(BaseDocumentParser):
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    def parse(self, file_path: str, **kwargs) -> dict:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            sheets_data[sheet_name] = rows
        wb.close()

        text_parts = []
        for name, rows in sheets_data.items():
            text_parts.append(f"Sheet: {name}")
            for row in rows:
                text_parts.append("\t".join(row))

        return {
            "content": "\n".join(text_parts),
            "sheets": sheets_data,
            "metadata": {"format": "xlsx", "sheet_count": len(sheets_data)},
        }
